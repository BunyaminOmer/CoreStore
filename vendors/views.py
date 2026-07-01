from decimal import Decimal, InvalidOperation
from datetime import timedelta
from io import BytesIO
import mimetypes
from pathlib import Path
import unicodedata
import urllib.error
import urllib.parse
import urllib.request

from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Count
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook

from .models import (
    AdPlacementRequest,
    SponsoredProductCampaign,
    SupportServiceOrder,
    SupportServicePackage,
    Vendor,
    VendorApplication,
    VendorSubscription,
    VendorSubscriptionPlan,
)
from .forms import (
    AdPlacementRequestForm,
    BulkProductUploadForm,
    BusinessProfileForm,
    SponsoredProductCampaignForm,
    SupportServiceOrderForm,
    VendorApplicationForm,
    VendorProductForm,
)
from store.models import (
    CargoStation,
    Category,
    Order,
    OrderItem,
    Product,
    ProductMedia,
    Shipment,
    ShipmentEvent,
    ShippingCompany,
)

def is_approved_vendor(user):
    return user.is_authenticated and user.is_vendor and hasattr(user, 'vendor_profile') and user.vendor_profile.is_approved


def require_business_profile(request, vendor):
    if vendor.has_business_profile:
        return None
    messages.warning(request, 'Abonelik, reklam ve hizmet paketlerinden önce işletme profilinizi tamamlayın.')
    return redirect('vendors:business_profile')


def latest_vendor_application(user):
    return user.vendor_applications.order_by('-created_at').first()


def get_corejet_company():
    company, _ = ShippingCompany.objects.get_or_create(
        code='CJ',
        defaults={
            'name': 'CoreJet',
            'support_phone': '0850 255 00 00',
            'tracking_url': '',
            'is_active': True,
        },
    )
    return company


BULK_COLUMN_ALIASES = {
    'name': {'urun_adi', 'urun', 'product_name', 'name', 'baslik'},
    'category': {'kategori', 'kategori_slug', 'kategori_id', 'category', 'category_slug', 'category_id'},
    'description': {'aciklama', 'description', 'detay'},
    'price': {'fiyat', 'satis_fiyati', 'price', 'normal_fiyat'},
    'supplier_price': {'tedarik_fiyati', 'alis_fiyati', 'supplier_price', 'maliyet'},
    'margin_rate': {'kar_marji', 'margin_rate', 'kar_orani'},
    'pricing_note': {'fiyat_bilgisi', 'pricing_note', 'fiyat_notu'},
    'discount_price': {'indirimli_fiyat', 'discount_price', 'kampanya_fiyati'},
    'stock': {'stok', 'stock', 'adet'},
    'is_active': {'aktif', 'is_active', 'yayinda'},
    'image_url': {'gorsel_url', 'resim_url', 'image_url', 'urun_gorseli'},
    'features': {'ozellikler', 'features', 'urun_ozellikleri'},
    'technical_info': {'teknik_bilgi', 'technical_info', 'teknik_ozellikler'},
    'brand': {'marka', 'brand'},
    'product_code': {'urun_kodu', 'product_code', 'sku'},
    'gtin': {'gtin', 'barkod', 'barcode'},
    'source_url': {'kaynak_url', 'source_url', 'urun_linki'},
}


def normalize_bulk_header(value):
    value = str(value or '').strip().lower()
    translation = str.maketrans('çğıöşüİ', 'cgiosui')
    value = value.translate(translation)
    value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    normalized = []
    for char in value:
        normalized.append(char if char.isalnum() else '_')
    return '_'.join(part for part in ''.join(normalized).split('_') if part)


def build_bulk_column_map(header_row):
    column_map = {}
    for index, value in enumerate(header_row):
        normalized = normalize_bulk_header(value)
        for target, aliases in BULK_COLUMN_ALIASES.items():
            if normalized in aliases:
                column_map[target] = index
    return column_map


def bulk_cell(row, column_map, key, default=''):
    index = column_map.get(key)
    if index is None or index >= len(row):
        return default
    value = row[index]
    if value is None:
        return default
    if isinstance(value, str):
        return value.strip()
    return value


def parse_bulk_decimal(value, field_name, required=True):
    if value in (None, ''):
        if required:
            raise ValueError(f'{field_name} boş bırakılamaz.')
        return None

    if isinstance(value, Decimal):
        return value

    text = str(value).strip().replace('₺', '').replace('TL', '').replace('tl', '').replace(' ', '')
    if ',' in text and '.' in text:
        if text.rfind(',') > text.rfind('.'):
            text = text.replace('.', '').replace(',', '.')
        else:
            text = text.replace(',', '')
    elif ',' in text:
        text = text.replace(',', '.')

    try:
        amount = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f'{field_name} geçerli bir sayı değil.') from exc

    if amount < 0:
        raise ValueError(f'{field_name} negatif olamaz.')
    return amount


def parse_bulk_stock(value):
    if value in (None, ''):
        raise ValueError('Stok boş bırakılamaz.')
    try:
        stock = int(float(str(value).replace(',', '.')))
    except ValueError as exc:
        raise ValueError('Stok geçerli bir tam sayı olmalıdır.') from exc
    if stock < 0:
        raise ValueError('Stok negatif olamaz.')
    return stock


def parse_bulk_bool(value, default=True):
    if value in (None, ''):
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in {'1', 'true', 'evet', 'aktif', 'yayinda', 'yayında', 'yes', 'y'}


def build_bulk_description(row, column_map):
    sections = []
    description = str(bulk_cell(row, column_map, 'description', '') or '').strip()
    if description:
        sections.append(description)

    details = [
        ('Ürün Özellikleri', bulk_cell(row, column_map, 'features', '')),
        ('Teknik Bilgi', bulk_cell(row, column_map, 'technical_info', '')),
    ]
    for title, value in details:
        value = str(value or '').strip()
        if value:
            sections.append(f'{title}\n{value}')

    pricing_lines = []
    supplier_price = bulk_cell(row, column_map, 'supplier_price', '')
    margin_rate = bulk_cell(row, column_map, 'margin_rate', '')
    pricing_note = str(bulk_cell(row, column_map, 'pricing_note', '') or '').strip()
    if supplier_price not in (None, ''):
        pricing_lines.append(f'Tedarik fiyatı: {supplier_price} TL')
    if margin_rate not in (None, ''):
        try:
            margin_percent = Decimal(str(margin_rate)) * Decimal('100')
            pricing_lines.append(f'Kar marjı: %{margin_percent.quantize(Decimal("1"))}')
        except (InvalidOperation, ValueError):
            pricing_lines.append(f'Kar marjı: {margin_rate}')
    if pricing_note:
        pricing_lines.append(pricing_note)
    if pricing_lines:
        sections.append('Fiyat Bilgisi\n' + '\n'.join(pricing_lines))

    meta_lines = []
    for label, key in (
        ('Marka', 'brand'),
        ('Ürün Kodu', 'product_code'),
        ('GTIN', 'gtin'),
        ('Kaynak', 'source_url'),
    ):
        value = str(bulk_cell(row, column_map, key, '') or '').strip()
        if value:
            meta_lines.append(f'{label}: {value}')
    if meta_lines:
        sections.append('Kaynak ve Ürün Bilgileri\n' + '\n'.join(meta_lines))

    return '\n\n'.join(sections)


def download_bulk_image(image_url):
    image_url = str(image_url or '').strip()
    if not image_url:
        return None

    parsed = urllib.parse.urlparse(image_url)
    if parsed.scheme not in {'http', 'https'}:
        return None

    request = urllib.request.Request(
        image_url,
        headers={'User-Agent': 'Mozilla/5.0 CoreLogicStoreBulkImporter/1.0'},
    )
    with urllib.request.urlopen(request, timeout=15) as response:
        content_type = response.headers.get_content_type()
        if not content_type.startswith('image/'):
            return None
        content = response.read(6 * 1024 * 1024)
        if len(content) >= 6 * 1024 * 1024:
            raise ValueError('Görsel dosyası çok büyük.')

    extension = mimetypes.guess_extension(content_type) or '.jpg'
    if extension == '.jpe':
        extension = '.jpg'
    filename_base = slugify(Path(parsed.path).stem or 'urun-gorseli')[:80] or 'urun-gorseli'
    return f'{filename_base}{extension}', ContentFile(content)


def find_bulk_category(value):
    if value in (None, ''):
        raise ValueError('Kategori boş bırakılamaz.')

    categories = Category.objects.filter(is_active=True)
    text = str(value).strip()
    if text.isdigit():
        category = categories.filter(id=int(text)).first()
        if category:
            return category

    category = categories.filter(slug__iexact=slugify(text)).first()
    if category:
        return category

    category = categories.filter(name__iexact=text).first()
    if category:
        return category

    raise ValueError(f'"{text}" kategorisi bulunamadı.')


def build_unique_product_slug(name):
    base_slug = slugify(name)[:250] or 'urun'
    slug = base_slug
    counter = 2
    while Product.objects.filter(slug=slug).exists():
        suffix = f'-{counter}'
        slug = f'{base_slug[:300 - len(suffix)]}{suffix}'
        counter += 1
    return slug


def process_bulk_product_upload(vendor, uploaded_file, update_existing=False, default_active=True):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    result = {
        'created': 0,
        'updated': 0,
        'errors': [],
        'items': [],
    }

    if not rows:
        result['errors'].append({'row': '-', 'message': 'Excel dosyası boş.'})
        return result

    column_map = build_bulk_column_map(rows[0])
    required_columns = {'name', 'category', 'price', 'stock'}
    missing_columns = required_columns - set(column_map)
    if missing_columns:
        readable = ', '.join(sorted(missing_columns))
        result['errors'].append({'row': 1, 'message': f'Eksik zorunlu kolon: {readable}.'})
        return result

    remaining_slots = max(vendor.product_limit - vendor.products.count(), 0)

    with transaction.atomic():
        for row_number, row in enumerate(rows[1:], start=2):
            if not any(value not in (None, '') for value in row):
                continue

            try:
                name = str(bulk_cell(row, column_map, 'name')).strip()
                if not name:
                    raise ValueError('Ürün adı boş bırakılamaz.')

                category = find_bulk_category(bulk_cell(row, column_map, 'category'))
                description = build_bulk_description(row, column_map)
                price = parse_bulk_decimal(bulk_cell(row, column_map, 'price'), 'Fiyat')
                discount_price = parse_bulk_decimal(
                    bulk_cell(row, column_map, 'discount_price', ''),
                    'İndirimli fiyat',
                    required=False,
                )
                stock = parse_bulk_stock(bulk_cell(row, column_map, 'stock'))
                is_active = parse_bulk_bool(
                    bulk_cell(row, column_map, 'is_active', ''),
                    default=default_active,
                )

                if discount_price is not None and discount_price >= price:
                    raise ValueError('İndirimli fiyat normal fiyattan düşük olmalıdır.')

                existing_product = Product.objects.filter(vendor=vendor, name__iexact=name).first()
                if existing_product and not update_existing:
                    raise ValueError('Bu isimde ürün zaten var. Güncelleme seçeneğini açın veya ürün adını değiştirin.')

                if existing_product:
                    product = existing_product
                    action = 'updated'
                else:
                    if result['created'] >= remaining_slots:
                        raise ValueError('Ürün ekleme limitiniz doldu.')
                    product = Product(vendor=vendor, slug=build_unique_product_slug(name))
                    action = 'created'

                product.name = name
                product.category = category
                product.description = description
                product.price = price
                product.discount_price = discount_price
                product.stock = stock
                product.is_active = is_active
                image_url = bulk_cell(row, column_map, 'image_url', '')
                if image_url and (not product.image or update_existing):
                    try:
                        image_payload = download_bulk_image(image_url)
                    except (urllib.error.URLError, TimeoutError, ValueError) as exc:
                        image_payload = None
                        result['errors'].append({
                            'row': row_number,
                            'message': f'Görsel indirilemedi, ürün görselsiz aktarıldı: {exc}',
                        })
                    if image_payload:
                        filename, content = image_payload
                        product.image.save(filename, content, save=False)
                product.save()

                result[action] += 1
                result['items'].append({
                    'row': row_number,
                    'name': product.name,
                    'category': category.name,
                    'action': 'Güncellendi' if action == 'updated' else 'Eklendi',
                })
            except ValueError as exc:
                result['errors'].append({'row': row_number, 'message': str(exc)})

    return result

@login_required
def vendor_portal_view(request):
    if is_approved_vendor(request.user):
        return redirect('vendors:dashboard')

    application = latest_vendor_application(request.user)
    if application:
        if application.status == VendorApplication.Status.PENDING:
            return redirect('vendors:pending')
        if application.status == VendorApplication.Status.APPROVED and hasattr(request.user, 'vendor_profile'):
            return redirect('vendors:dashboard')

    return redirect('vendors:apply')


@login_required
def vendor_apply_view(request):
    if is_approved_vendor(request.user):
        messages.success(request, 'Zaten onaylı bir satıcısınız.')
        return redirect('vendors:dashboard')

    app = latest_vendor_application(request.user)
    if app:
        if app.status == VendorApplication.Status.PENDING:
            return redirect('vendors:pending')
        if app.status == VendorApplication.Status.APPROVED:
            messages.info(request, 'Başvurunuz onaylandı. Satıcı paneliniz hazırlanıyor.')
            return redirect('vendors:pending')
            
    if request.method == 'POST':
        form = VendorApplicationForm(request.POST)
        if form.is_valid():
            application = form.save(commit=False)
            application.user = request.user
            application.save()
            messages.success(request, 'Satıcı başvurunuz alındı.')
            return redirect('vendors:pending')
    else:
        form = VendorApplicationForm()
        
    return render(request, 'vendors/apply.html', {'form': form})

@login_required
def vendor_pending_view(request):
    if is_approved_vendor(request.user):
        return redirect('vendors:dashboard')

    app = latest_vendor_application(request.user)
    if not app:
        return redirect('vendors:apply')
    return render(request, 'vendors/pending.html', {'application': app})

@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_dashboard_view(request):
    from django.db.models import F
    import json
    
    vendor = request.user.vendor_profile
    active_subscription = vendor.active_subscription
    business_profile = getattr(vendor, 'business_profile', None)
    
    # Stats
    product_count = vendor.products.count()
    active_orders = OrderItem.objects.filter(product__vendor=vendor, order__status__in=['pending', 'confirmed', 'shipped'])
    active_orders_count = active_orders.values('order').distinct().count()
    
    # Total sales
    total_sales = OrderItem.objects.filter(product__vendor=vendor, order__status='delivered').aggregate(
        total=Sum(F('price') * F('quantity'))
    )['total'] or 0
    
    # Recent orders
    recent_orders = OrderItem.objects.filter(product__vendor=vendor).order_by('-order__created_at')[:5]
    
    # Low stock products
    low_stock_products = vendor.products.filter(stock__lt=5, is_active=True).order_by('stock')
    
    # Sales Chart Data (last 7 days)
    today = timezone.now().date()
    sales_labels = []
    sales_data = []
    
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        day_sales = OrderItem.objects.filter(
            product__vendor=vendor,
            order__status='delivered',
            order__created_at__date=day
        ).aggregate(total=Sum(F('price') * F('quantity')))['total'] or 0
        
        sales_labels.append(day.strftime("%d %b"))
        sales_data.append(float(day_sales))
    
    context = {
        'vendor': vendor,
        'business_profile': business_profile,
        'active_subscription': active_subscription,
        'pending_subscription_count': vendor.subscriptions.filter(status=VendorSubscription.Status.PENDING).count(),
        'pending_campaign_count': vendor.sponsored_campaigns.filter(status=SponsoredProductCampaign.Status.PENDING).count(),
        'active_campaign_count': vendor.sponsored_campaigns.filter(status=SponsoredProductCampaign.Status.ACTIVE).count(),
        'pending_ad_count': vendor.ad_requests.filter(status=AdPlacementRequest.Status.PENDING).count(),
        'open_service_count': vendor.service_orders.exclude(status=SupportServiceOrder.Status.COMPLETED).count(),
        'product_count': product_count,
        'active_orders_count': active_orders_count,
        'total_sales': total_sales,
        'recent_orders': recent_orders,
        'low_stock_products': low_stock_products,
        'sales_labels': json.dumps(sales_labels),
        'sales_data': json.dumps(sales_data),
    }
    return render(request, 'vendors/dashboard.html', context)


@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_business_profile_view(request):
    vendor = request.user.vendor_profile
    profile = getattr(vendor, 'business_profile', None)

    if request.method == 'POST':
        form = BusinessProfileForm(request.POST, instance=profile)
        if form.is_valid():
            profile = form.save(commit=False)
            profile.vendor = vendor
            profile.save()
            messages.success(request, 'İşletme profiliniz kaydedildi. Artık paket ve reklam araçlarını kullanabilirsiniz.')
            return redirect('vendors:subscriptions')
    else:
        initial = {
            'legal_name': vendor.store_name,
            'tax_number': vendor.tax_number,
            'phone': vendor.phone,
            'billing_email': request.user.email,
        }
        form = BusinessProfileForm(instance=profile, initial=initial)

    return render(request, 'vendors/business_profile.html', {
        'vendor': vendor,
        'profile': profile,
        'form': form,
    })


@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_subscription_view(request):
    vendor = request.user.vendor_profile
    redirect_response = require_business_profile(request, vendor)
    if redirect_response:
        return redirect_response

    plans = VendorSubscriptionPlan.objects.filter(is_active=True).order_by('display_order', 'monthly_price')
    active_subscription = vendor.active_subscription
    subscriptions = vendor.subscriptions.select_related('plan')[:8]

    if request.method == 'POST':
        plan = get_object_or_404(plans, id=request.POST.get('plan_id'))
        starts_at = timezone.now()
        status = VendorSubscription.Status.ACTIVE if plan.monthly_price == 0 else VendorSubscription.Status.PENDING
        subscription = VendorSubscription.objects.create(
            vendor=vendor,
            plan=plan,
            status=status,
            starts_at=starts_at,
            ends_at=starts_at + timedelta(days=30),
            payment_note='Demo mod: ödeme entegrasyonu bağlanınca tahsilat adımı eklenecek.',
        )
        if status == VendorSubscription.Status.ACTIVE:
            vendor.subscriptions.exclude(id=subscription.id).filter(
                status=VendorSubscription.Status.ACTIVE,
            ).update(status=VendorSubscription.Status.CANCELLED)
            vendor.product_limit = plan.product_limit
            vendor.save(update_fields=['product_limit', 'updated_at'])
            messages.success(request, f'{plan.name} paketi aktif edildi.')
        else:
            messages.success(request, f'{plan.name} paketi için talebiniz alındı. Yönetici onayından sonra aktifleşecek.')
        return redirect('vendors:subscriptions')

    return render(request, 'vendors/subscriptions.html', {
        'vendor': vendor,
        'plans': plans,
        'active_subscription': active_subscription,
        'subscriptions': subscriptions,
    })


@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_sponsorships_view(request):
    vendor = request.user.vendor_profile
    redirect_response = require_business_profile(request, vendor)
    if redirect_response:
        return redirect_response
    active_subscription = vendor.active_subscription

    if not active_subscription:
        messages.warning(request, 'Sponsorlu ürün talebi oluşturmak için önce aktif bir satıcı paketi seçin.')
        return redirect('vendors:subscriptions')

    if request.method == 'POST':
        form = SponsoredProductCampaignForm(request.POST, vendor=vendor)
        if form.is_valid():
            used_quota = vendor.sponsored_campaigns.filter(
                status__in=[
                    SponsoredProductCampaign.Status.PENDING,
                    SponsoredProductCampaign.Status.ACTIVE,
                ],
                starts_at__gte=active_subscription.starts_at,
            ).count()
            if used_quota >= active_subscription.plan.sponsored_product_quota:
                messages.error(request, 'Bu paketteki sponsorlu ürün hakkınız doldu. Paketinizi yükseltebilir veya mevcut taleplerin bitmesini bekleyebilirsiniz.')
                return redirect('vendors:sponsorships')
            campaign = form.save(commit=False)
            campaign.vendor = vendor
            campaign.status = SponsoredProductCampaign.Status.PENDING
            campaign.save()
            messages.success(request, 'Sponsorlu ürün talebiniz alındı. Onaylandığında ilgili vitrinde gösterilecek.')
            return redirect('vendors:sponsorships')
    else:
        form = SponsoredProductCampaignForm(vendor=vendor)

    campaigns = vendor.sponsored_campaigns.select_related('product').order_by('-created_at')
    return render(request, 'vendors/sponsorships.html', {
        'vendor': vendor,
        'form': form,
        'campaigns': campaigns,
        'active_subscription': active_subscription,
    })


@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_ads_view(request):
    vendor = request.user.vendor_profile
    redirect_response = require_business_profile(request, vendor)
    if redirect_response:
        return redirect_response
    active_subscription = vendor.active_subscription

    if not active_subscription:
        messages.warning(request, 'Reklam alanı talebi oluşturmak için önce aktif bir satıcı paketi seçin.')
        return redirect('vendors:subscriptions')

    if request.method == 'POST':
        form = AdPlacementRequestForm(request.POST, request.FILES)
        if form.is_valid():
            ad_request = form.save(commit=False)
            ad_request.vendor = vendor
            ad_request.status = AdPlacementRequest.Status.PENDING
            ad_request.save()
            messages.success(request, 'Reklam alanı talebiniz alındı. Yönetici fiyat ve yayın durumunu panelden netleştirecek.')
            return redirect('vendors:ads')
    else:
        form = AdPlacementRequestForm()

    ad_requests = vendor.ad_requests.order_by('-created_at')
    return render(request, 'vendors/ads.html', {
        'vendor': vendor,
        'form': form,
        'ad_requests': ad_requests,
        'active_subscription': active_subscription,
    })


@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_services_view(request):
    vendor = request.user.vendor_profile
    redirect_response = require_business_profile(request, vendor)
    if redirect_response:
        return redirect_response

    packages = SupportServicePackage.objects.filter(is_active=True).order_by('display_order', 'price')
    if request.method == 'POST':
        package = get_object_or_404(packages, id=request.POST.get('package_id'))
        form = SupportServiceOrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.vendor = vendor
            order.package = package
            order.save()
            messages.success(request, f'{package.name} talebiniz alındı. Destek ekibi sizinle iletişime geçecek.')
            return redirect('vendors:services')
    else:
        form = SupportServiceOrderForm()

    service_orders = vendor.service_orders.select_related('package').order_by('-created_at')
    return render(request, 'vendors/services.html', {
        'vendor': vendor,
        'packages': packages,
        'form': form,
        'service_orders': service_orders,
    })

@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_product_list_view(request):
    vendor = request.user.vendor_profile
    products = vendor.products.all().order_by('-created_at')
    return render(request, 'vendors/product_list.html', {'products': products, 'vendor': vendor})

@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_product_add_view(request):
    vendor = request.user.vendor_profile
    if not vendor.can_add_product:
        messages.error(request, 'Ürün ekleme limitinize ulaştınız.')
        return redirect('vendors:product_list')
        
    if request.method == 'POST':
        form = VendorProductForm(request.POST, request.FILES, vendor=vendor)
        if form.is_valid():
            product = form.save(commit=False)
            product.vendor = vendor
            product.save()
            form.save_media(product)
            form.save_variants(product)
            messages.success(request, 'Ürün başarıyla eklendi.')
            return redirect('vendors:product_list')
    else:
        form = VendorProductForm(vendor=vendor)
        
    return render(request, 'vendors/product_form.html', {'form': form, 'title': 'Yeni Ürün Ekle'})

@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_product_edit_view(request, pk):
    vendor = request.user.vendor_profile
    product = get_object_or_404(Product.objects.prefetch_related('media'), pk=pk, vendor=vendor)
    
    if request.method == 'POST':
        form = VendorProductForm(request.POST, request.FILES, instance=product, vendor=vendor)
        if form.is_valid():
            product = form.save()
            delete_ids = request.POST.getlist('delete_media')
            if delete_ids:
                ProductMedia.objects.filter(product=product, id__in=delete_ids).delete()
            form.save_media(product)
            form.save_variants(product)
            messages.success(request, 'Ürün güncellendi.')
            return redirect('vendors:product_list')
    else:
        form = VendorProductForm(instance=product, vendor=vendor)
        
    return render(request, 'vendors/product_form.html', {'form': form, 'title': 'Ürün Düzenle', 'product': product})

@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_product_delete_view(request, pk):
    vendor = request.user.vendor_profile
    product = get_object_or_404(Product, pk=pk, vendor=vendor)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Ürün silindi.')
    return redirect('vendors:product_list')

@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_orders_view(request):
    vendor = request.user.vendor_profile
    status_filter = request.GET.get('status')
    
    order_items = OrderItem.objects.filter(product__vendor=vendor).select_related(
        'order',
        'order__user',
        'order__shipment',
        'order__shipment__company',
        'order__shipment__current_station',
        'product',
    )
    if status_filter:
        order_items = order_items.filter(order__status=status_filter)
        
    order_items = order_items.order_by('-order__created_at')
    
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        new_status = request.POST.get('status')
        order = get_object_or_404(Order, id=order_id)
        if order.items.filter(product__vendor=vendor).exists():
            if new_status:
                order.status = new_status
                order.save(update_fields=['status', 'updated_at'])

            shipment_status = request.POST.get('shipment_status') or Shipment.Status.CREATED
            if shipment_status not in Shipment.Status.values:
                shipment_status = Shipment.Status.CREATED
            tracking_number = (request.POST.get('tracking_number') or '').strip()
            station_id = request.POST.get('current_station')
            shipment_note = (request.POST.get('shipment_note') or '').strip()
            company = get_corejet_company()

            if tracking_number and Shipment.objects.filter(tracking_number=tracking_number).exclude(order=order).exists():
                messages.error(request, 'Bu takip numarası başka bir siparişte kullanılıyor.')
                return redirect('vendors:orders')

            shipment, created = Shipment.objects.get_or_create(
                order=order,
                defaults={
                    'company': company,
                    'tracking_number': tracking_number,
                    'status': shipment_status,
                    'note': shipment_note,
                },
            )

            previous_status = shipment.status
            previous_station_id = shipment.current_station_id
            if tracking_number:
                shipment.tracking_number = tracking_number
            shipment.company = company
            shipment.status = shipment_status
            shipment.note = shipment_note
            shipment.current_station = CargoStation.objects.filter(
                id=station_id,
                company=company,
                is_active=True,
            ).first() if station_id else None
            shipment.save()

            if created or previous_status != shipment.status or previous_station_id != shipment.current_station_id or shipment_note:
                ShipmentEvent.objects.create(
                    shipment=shipment,
                    station=shipment.current_station,
                    status=shipment.status,
                    description=shipment_note or 'Kargo bilgisi güncellendi.',
                )

            if shipment.status == Shipment.Status.DELIVERED:
                order.status = Order.Status.DELIVERED
                order.save(update_fields=['status', 'updated_at'])
            elif shipment.status in {Shipment.Status.PICKED_UP, Shipment.Status.IN_TRANSIT, Shipment.Status.AT_STATION, Shipment.Status.OUT_FOR_DELIVERY}:
                order.status = Order.Status.SHIPPED
                order.save(update_fields=['status', 'updated_at'])

            messages.success(request, f'Sipariş #{order.id} kargo bilgisi güncellendi.')
            return redirect('vendors:orders')

    corejet = get_corejet_company()
    cargo_stations = CargoStation.objects.filter(company=corejet, is_active=True)

    return render(request, 'vendors/orders.html', {
        'order_items': order_items,
        'cargo_stations': cargo_stations,
        'shipment_status_choices': Shipment.Status.choices,
    })

@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_bulk_upload_view(request):
    vendor = request.user.vendor_profile
    result = None

    if request.method == 'POST':
        form = BulkProductUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                result = process_bulk_product_upload(
                    vendor=vendor,
                    uploaded_file=form.cleaned_data['file'],
                    update_existing=form.cleaned_data['update_existing'],
                    default_active=form.cleaned_data['activate_products'],
                )
                if result['created'] or result['updated']:
                    messages.success(
                        request,
                        f"{result['created']} ürün eklendi, {result['updated']} ürün güncellendi.",
                    )
                if result['errors']:
                    messages.warning(request, f"{len(result['errors'])} satır içe aktarılamadı.")
            except Exception as exc:
                messages.error(request, f'Excel dosyası okunamadı: {exc}')
    else:
        form = BulkProductUploadForm()

    return render(request, 'vendors/bulk_upload.html', {
        'form': form,
        'result': result,
        'vendor': vendor,
    })


@user_passes_test(is_approved_vendor, login_url='/vendor/')
def vendor_bulk_upload_template_view(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Toplu Ürün'
    headers = [
        'Ürün Adı',
        'Kategori',
        'Açıklama',
        'Fiyat',
        'Tedarik Fiyatı',
        'Kar Marjı',
        'Fiyat Bilgisi',
        'İndirimli Fiyat',
        'Stok',
        'Aktif',
        'Görsel URL',
        'Özellikler',
        'Teknik Bilgi',
        'Marka',
        'Ürün Kodu',
        'GTIN',
        'Kaynak URL',
    ]
    sheet.append(headers)
    sheet.append([
        'Örnek Bluetooth Kulaklık',
        'Elektronik',
        'Kısa ürün açıklaması',
        '1299.90',
        '899.90',
        '0.50',
        'Tedarik fiyatı ve kar marjı açıklamaya eklenir.',
        '999.90',
        '25',
        'evet',
        'https://example.com/gorsel.jpg',
        'Bluetooth 5.3\nAktif gürültü azaltma',
        'Pil: 30 saat',
        'Örnek Marka',
        'SKU-001',
        '',
        'https://example.com/urun',
    ])
    sheet.append([
        'Örnek Telefon Kılıfı',
        'Aksesuar',
        'Kategori adı, slug veya kategori id yazabilirsiniz.',
        '249.90',
        '178.50',
        '0.40',
        '',
        '',
        '100',
        'evet',
        '',
        'Esnek silikon\nTam koruma',
        '',
        '',
        '',
        '',
        '',
    ])

    widths = [28, 22, 48, 14, 16, 12, 36, 18, 10, 12, 42, 36, 36, 18, 18, 16, 42]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="toplu-urun-sablonu.xlsx"'
    return response
